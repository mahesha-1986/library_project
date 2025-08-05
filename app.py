import io
from flask import Flask, jsonify, make_response, render_template, request, redirect, url_for, session, flash, send_file
import json
import logging
from openpyxl import Workbook
import pandas as pd
from bson import ObjectId
from DB.connection import db
from datetime import datetime
# Uncomment this line if you want to import books from an Excel file, white initial setup of the application
#import services.import_books_from_excel 
from services import Read_DepartmentCodes
import os
import uuid
from werkzeug.utils import secure_filename

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

app.secret_key = 'Student_Barcode_Data'

print("APP Started")
books_collection = db['books']
issued_books = db['issued_books']

# Configure upload settings
UPLOAD_FOLDER = 'static/clear_history_excel'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Create upload directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Load student data
folder_path = "students_data_JSON"

student_data = {}

# Loop through all JSON files in the folder
for file_name in os.listdir(folder_path):
    if file_name.endswith(".json"):  # Only process JSON files
        file_path = os.path.join(folder_path, file_name)
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            student_data.update(data) 

@app.route('/')
def home():
    try:
        All_issued_books_history = issued_books.find()
        All_issued_books_history = list(All_issued_books_history)
         # Reverse the order to show the latest issued books first
        All_issued_books_history.reverse()
        for b in All_issued_books_history:
            b['_id'] = str(b['_id'])
            
        logger.info("All issued books history loaded successfully.")

    except Exception as e:
        logger.error("Error loading home page", 500)


    return render_template('book_system.html',Books_history= All_issued_books_history)


@app.route("/search_statistics", methods=["GET", "POST"])
def search_statistics():
    try:
        total_books = None
        issued_books_count = None
        available_books = None
        header = None

        if request.method == "POST":
            query = request.form.get("query", "").strip()
            search_type = request.form.get("search_type", "").strip().lower()

            if not query or not search_type:
                return render_template(
                    'statistics.html',
                    error="Please enter both query and search type."
                )
            
           


            # Prepare query value based on search type
            if search_type == "department code":
                query_value = query.upper()
                search_type = "department_code"
            elif search_type == "title":
                query_value = query.upper()
                search_type = "title"
            else:
                query_value = query.title()

            header = f"{search_type.title() if search_type != 'department_code' else 'Department Code'}: {query_value}"

            # Prepare MongoDB filter
            filter_query = {search_type: query_value}
            filter_query_issued = {"book." + search_type: query_value, "status": "issued"}

            # Count total books
            total_books = books_collection.count_documents(filter_query)

            # Count issued books
            issued_books_count = issued_books.count_documents(filter_query_issued)
            print("total", total_books, "issued", issued_books_count)
            available_books = total_books - issued_books_count

        # Always render the page
        return render_template(
            'statistics.html',
            total_books=total_books,
            issued_books=issued_books_count,
            available_books=available_books,
            header=header,
        )

    except Exception as e:
        logger.error(f"Error in search_statistics: {e}")
        return jsonify({"error": "Error processing search request"}), 500
    

@app.route('/clear_history')
def clear_history():
    try:
        # Get all issued books history before clearing
        all_issued_books = list(issued_books.find())
        
        if all_issued_books:
            # Extract date range from database records
            from datetime import datetime
            
            # Get all issued_at dates and convert them to datetime objects
            issued_dates = []
            for book in all_issued_books:
                if book.get('issued_at'):
                    try:
                        # Parse the issued_at string to datetime
                        date_obj = datetime.strptime(book['issued_at'], '%Y-%m-%d %H:%M:%S')
                        issued_dates.append(date_obj)
                    except ValueError:
                        # If parsing fails, skip this record
                        continue
            
            if issued_dates:
                # Find the earliest and latest dates
                earliest_date = min(issued_dates)
                latest_date = max(issued_dates)
                
                # Format dates as date_month_year
                first_date_str = earliest_date.strftime('%d_%m_%Y')
                last_date_str = latest_date.strftime('%d_%m_%Y')
                
                filename = f"Issue_return_history_{first_date_str}_to_{last_date_str}.xlsx"
            else:
                # Fallback to timestamp if no valid dates found
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"Issue_return_history_{timestamp}.xlsx"
            
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            fileCreated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            # Prepare data for Excel
            excel_data = []
            for book in all_issued_books:
                excel_data.append({
                    'Student Name': book.get('student', {}).get('studentName', ''),
                    'Roll No': book.get('student', {}).get('rollno', ''),
                    'Section': book.get('student', {}).get('section', ''),
                    'Book Title': book.get('book', {}).get('title', ''),
                    'Book Author': book.get('book', {}).get('author', ''),
                    'Barcode': book.get('book', {}).get('barcode', ''),
                    'Department': book.get('book', {}).get('department', ''),
                    'Status': book.get('status', ''),
                    'Issued At': book.get('issued_at', ''),
                    'Returned At': book.get('returned_at', ''),
                    'Created Date' : fileCreated

                })
            
            # Create Excel file with proper column formatting
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Cleared Book History"
            
            # Define headers
            headers = [
                'Student Name', 'Roll No', 'Section', 'Book Title', 'Book Author',
                'Barcode', 'Department', 'Status', 'Issued At', 'Returned At',
                'Created Date'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for row_idx, book in enumerate(all_issued_books, 2):
                ws.cell(row=row_idx, column=1, value=book.get('student', {}).get('studentName', ''))
                ws.cell(row=row_idx, column=2, value=book.get('student', {}).get('rollno', ''))
                ws.cell(row=row_idx, column=3, value=book.get('student', {}).get('section', ''))
                ws.cell(row=row_idx, column=4, value=book.get('book', {}).get('title', ''))
                ws.cell(row=row_idx, column=5, value=book.get('book', {}).get('author', ''))
                ws.cell(row=row_idx, column=6, value=book.get('book', {}).get('barcode', ''))
                ws.cell(row=row_idx, column=7, value=book.get('book', {}).get('department','').upper())
                ws.cell(row=row_idx, column=8, value=book.get('status', '').upper())
                ws.cell(row=row_idx, column=9, value=book.get('issued_at', ''))
                ws.cell(row=row_idx, column=10, value=book.get('returned_at', ''))
                ws.cell(row=row_idx, column=11, value=fileCreated)
            
            # Auto-adjust column widths based on content
            for col in range(1, len(headers) + 1):
                max_length = 0
                column_letter = get_column_letter(col)
                
                # Check header length
                header_length = len(headers[col-1])
                max_length = max(max_length, header_length)
                
                # Check all data in this column
                for row_idx in range(2, len(all_issued_books) + 2):
                    cell_value = ws.cell(row=row_idx, column=col).value
                    if cell_value:
                        cell_length = len(str(cell_value))
                        max_length = max(max_length, cell_length)
                
                # Set optimal column width with padding
                if max_length > 0:
                    # Add padding based on content length
                    if max_length <= 10:
                        optimal_width = max_length + 8  # Extra padding for short content
                    elif max_length <= 20:
                        optimal_width = max_length + 6  # Good padding for medium content
                    elif max_length <= 30:
                        optimal_width = max_length + 4  # Adequate padding for longer content
                    else:
                        optimal_width = max_length + 2  # Minimal padding for very long content
                    
                    # Apply constraints (min 15, max 80)
                    final_width = max(15, min(optimal_width, 80))
                    ws.column_dimensions[column_letter].width = final_width
                else:
                    # Default width for empty columns
                    ws.column_dimensions[column_letter].width = 20
            
            # Format data rows
            data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(all_issued_books) + 2):  # +2 for header row
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # Save the workbook
            wb.save(file_path)
            
            # Clear the history
            issued_books.delete_many({})
            
            flash(f'All book transaction history cleared successfully. Excel backup saved as "{filename}"', 'success')
        else:
            flash('No transaction history to clear.', 'info')
            
    except Exception as e:
        logger.error(f"Error clearing history: {e}")
        flash(f'Error clearing history: {str(e)}', 'error')
    
    return redirect(url_for('home'))

@app.route('/view_excel_files')
def view_excel_files():
    files = []
    if os.path.exists(UPLOAD_FOLDER):
        for filename in os.listdir(UPLOAD_FOLDER):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(UPLOAD_FOLDER, filename)
                file_size = os.path.getsize(file_path)
                file_date = datetime.fromtimestamp(os.path.getctime(file_path))
                files.append({
                    'name': filename,
                    'size': file_size,
                    'date': file_date.strftime('%Y-%m-%d %H:%M:%S')
                })
    
    # Sort files by date (newest first)
    files.sort(key=lambda x: x['date'], reverse=True)
    
    return render_template('view_excel_files.html', files=files)

@app.route('/download_excel_file/<filename>')
def download_excel_file(filename):
    try:
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            flash('File not found', 'error')
            return redirect(url_for('view_excel_files'))
    except Exception as e:
        flash(f'Error downloading file: {str(e)}', 'error')
        return redirect(url_for('view_excel_files'))

@app.route('/delete_excel_file/<filename>')
def delete_excel_file(filename):
    try:
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            flash(f'File "{filename}" deleted successfully!', 'success')
        else:
            flash('File not found', 'error')
    except Exception as e:
        flash(f'Error deleting file: {str(e)}', 'error')
    
    return redirect(url_for('view_excel_files'))

@app.route('/clear_all_excel_files')
def clear_all_excel_files():
    try:
        if os.path.exists(UPLOAD_FOLDER):
            for filename in os.listdir(UPLOAD_FOLDER):
                if filename.endswith('.xlsx'):
                    file_path = os.path.join(UPLOAD_FOLDER, filename)
                    os.remove(file_path)
            flash('All Excel files cleared successfully!', 'success')
        else:
            flash('No files to clear', 'info')
    except Exception as e:
        flash(f'Error clearing files: {str(e)}', 'error')
    
    return redirect(url_for('view_excel_files'))

@app.route('/deleted_data')
def deleted_data():
    files = []
    excel_folder = 'static/clear_history_excel'
    
    if os.path.exists(excel_folder):
        for filename in os.listdir(excel_folder):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(excel_folder, filename)
                file_size = os.path.getsize(file_path)
                file_date = datetime.fromtimestamp(os.path.getctime(file_path))
                files.append({
                    'filename': filename,
                    'size': file_size,
                    'created': file_date.strftime('%Y-%m-%d %H:%M:%S')
                })
    
    # Sort files by date (newest first)
    files.sort(key=lambda x: x['created'], reverse=True)
    
    return render_template('deleted_data.html', files=files)

@app.route('/download_deleted_data/<filename>')
def download_deleted_data(filename):
    try:
        file_path = os.path.join('static/clear_history_excel', filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            flash('File not found', 'error')
            return redirect(url_for('deleted_data'))
    except Exception as e:
        flash(f'Error downloading file: {str(e)}', 'error')
        return redirect(url_for('deleted_data'))

@app.route('/delete_deleted_data/<filename>', methods=['POST'])
def delete_deleted_data(filename):
    try:
        file_path = os.path.join('static/clear_history_excel', filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            flash(f'File "{filename}" deleted successfully!', 'success')
        else:
            flash('File not found', 'error')
    except Exception as e:
        flash(f'Error deleting file: {str(e)}', 'error')
    return redirect(url_for('deleted_data'))

@app.route("/recommendation")
def recommendation():
    search_type = request.args.get("type", "Title").lower()
    query = request.args.get("query", "")
    
    
    try:
        if search_type == "department code":
            search_type = "department_code"

        cursor = books_collection.find(
                {search_type: {"$regex": query, "$options": "i"}},
                {search_type: 1, "_id": 0}
        )
        
        result_list = [doc[search_type] for doc in cursor]
        result_list = list(set(result_list))  # Convert to set and back to list to remove duplicates

        
        suggestions = [s for s in result_list if query.lower() in s.lower()]
        return jsonify(suggestions)

    except Exception as e:
        logger.error(f"Error fetching data for recommendations: {e}")
        return jsonify({"error": "Error fetching data"}), 500
   



@app.route('/returned_books')
def return_book():
    id = request.args.get('id')
    if not id:
        return "Book ID is required", 400

    try:
        issued_books.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'status': 'returned', 'returned_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}}
        )
        logger.info(f"Book with ID {id} returned successfully.")
    except Exception as e:
        logger.error(f"Error returning book with ID {id}: {e}")
        return "Error returning book", 500

    return redirect(url_for('home'))

@app.route('/all-books')
def all_books():
    try:
        departments = Read_DepartmentCodes()
        books_cursor = books_collection.find()
        books = [{**book, '_id': str(book['_id'])} for book in books_cursor]

        for book in books:
            book_history = issued_books.find_one({'book.barcode': book['barcode'], 'status': 'issued'})
           
            if not book_history:
                book['status'] = 'available'
            else:
                book['status'] = 'not available'
                book['student'] = book_history.get('student')
                book['issued_at'] = book_history.get('issued_at')

    except Exception as e:
        logger.error("Error fetching books: %s", e)
        return "Error fetching books", 500
    return render_template("all_books.html", Books=books, departments=departments, Book_Count=len(books))



@app.route('/issue-book', methods=['GET', 'POST'])
def index():
    result = None
    error = None
    student_books_count = None
    show_limit_popup = False
    
    if request.method == 'POST':
        key = request.form.get('rollno', '').strip().upper()
        if key in student_data:
            result = {'rollno': key, **student_data[key]}
            
            # Check if student has reached the 4-book limit
            student_books_count = issued_books.count_documents({
                'student.rollno': key,
                'status': 'issued'
            })
            
            if student_books_count >= 4:
                show_limit_popup = True
        else:
            error = f"No data found for roll number: {key}"
    
    return render_template('index.html', result=result, error=error, student_books_count=student_books_count, show_limit_popup=show_limit_popup)

@app.route('/view-books')
def view_books():
    key = request.args.get('rollno')
    if key:
        session['student'] = {'rollno': key, **student_data[key]}

    barcode = request.args.get('barcode')
    barcode = barcode.upper() if barcode else None
    barcode_result = None
    barcode_searched = None
    already_issued = None
    student_books_count = None

    try:
        if barcode:
            barcode_result = books_collection.find_one({'barcode': barcode})
            if barcode_result:
                # remove _id or convert it
                barcode_result['id'] = str(barcode_result.pop('_id'))
                session['book'] = barcode_result

                already_issued = issued_books.find_one({'book.barcode': barcode,'status': 'issued'})
                if already_issued:
                    # convert any ObjectId inside already_issued too
                    already_issued['_id'] = str(already_issued['_id'])

            else:
                barcode_searched = barcode

        # Get student info from session to show current book count
        student = session.get('student')
        if student:
            student_books_count = issued_books.count_documents({
                'student.rollno': student['rollno'],
                'status': 'issued'
            })

    except Exception as e:
        logger.error("Error in view_books route: %s", e)

    books = []
    try:
        books_cursor = books_collection.find()
        books = [{**book, '_id': str(book['_id'])} for book in books_cursor]
    except Exception as e:
        logger.error("Error fetching books: %s", e)

    # Get error message from session if exists
    book_limit_error = session.pop('book_limit_error', None)

    return render_template(
        'view_books.html',
        
        barcode_result=barcode_result,
        barcode_searched=barcode_searched,
        already_issued=already_issued,
        book_limit_error=book_limit_error,
        student_books_count=student_books_count
    )

@app.route('/Save')
def Save_To_DB():
    student = session.get('student')
    barcode_result = session.get('book')
    if not student or not barcode_result:
        return "Missing student or book information", 400

    # Check if student already has 4 books issued
    student_issued_books = issued_books.count_documents({
        'student.rollno': student['rollno'],
        'status': 'issued'
    })
    
    if student_issued_books >= 4:
        # Store error message in session for popup
        session['book_limit_error'] = f"Student {student['studentName']} (Roll No: {student['rollno']}) has already reached the maximum limit of 4 books. Please return some books before issuing new ones."
        return redirect(url_for('view_books'))

    issued_doc = {
        'student': student,
        'book': barcode_result,
        'status': 'issued',
        'issued_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # readable current date & time
    }

    try:
        issued_books.insert_one(issued_doc)
        logger.info("Book issued successfully!")
        # Clear any previous error messages
        session.pop('book_limit_error', None)
    except Exception as e:
        logger.error("Error while issuing book: %s", e)
        return "Error issuing book", 500

    return redirect(url_for('home'))


@app.route('/add-book', methods=['GET', 'POST'])
def add_book():
    print("Add Book Page")
    if request.method == 'POST':
        print("Add Book Page111")
        name = str(request.form['book_name'])
        dept = str(request.form['department'])
        author = str(request.form['author'])
        accession_number = request.form['accession_number']
        try:
            
            departments = Read_DepartmentCodes()
            
            if dept in departments:
                dept_code = departments[dept]
            else:
                logger.error(f"Invalid department: {dept}")
                return "Invalid department", 400
            barcode = dept_code +  str(accession_number)

            book_data = {
                'title': name.upper(),
                'barcode': barcode.upper(),
                'author': author.title(),
                'accession_number': accession_number,
                'department': dept.title(),
                'department_code': dept_code.upper(),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # readable current date & time
            }
            books_collection.insert_one(book_data)
            logger.info("Book added successfully.")
            flash(f"Book '{name}' added successfully!", "success")
            return redirect(url_for('all_books'))
 
        except Exception as e:
            logger.error(f"Error adding book: {e}")
            return "Error adding book", 500





@app.route('/lookup_barcode', methods=['GET'])
def lookup_barcode():
    barcode = request.args.get('barcode')
    if not barcode:
        return jsonify({"error": "No barcode provided"}), 400
    book = books_collection.find_one({'barcode': barcode})
    if not book:
        return jsonify({"error": "Book not found"}), 404
    book.pop('_id', None)
    return jsonify(book)

@app.route('/delete_entry/<id>')
def delete_entry(id):
    try:
        from bson import ObjectId
        result = issued_books.delete_one({'_id': ObjectId(id)})
        if result.deleted_count == 0:
            return 'Entry not found', 404
    except Exception as e:
        logger.error(f"Error deleting entry with ID {id}: {e}")
        return 'Error deleting entry', 500
    return redirect(url_for('home'))

@app.route("/delete_book/<accession_number>", methods=["DELETE"])
def delete_book(accession_number):
    try:
        print("Received accession_number:", accession_number, "Type:", type(accession_number))
        
        # Try to find the book with the original accession number
        book = books_collection.find_one({"accession_number": accession_number})
        print("Found book with original:", book)
        
        if not book:
            # Try with different data types to handle potential type mismatches
            if accession_number.isdigit():
                # Try as integer
                book = books_collection.find_one({"accession_number": int(accession_number)})
                print("Found book with int:", book)
                if book:
                    # Use the integer version for deletion
                    result = books_collection.delete_one({"accession_number": int(accession_number)})
                    if result.deleted_count > 0:
                        # Set flash message for success
                        flash(f"Book with accession number {accession_number} deleted successfully!", "success")
                        return jsonify({"success": True, "message": "Book deleted successfully."})
            else:
                # Try as string
                book = books_collection.find_one({"accession_number": str(accession_number)})
                print("Found book with str:", book)
                if book:
                    # Use the string version for deletion
                    result = books_collection.delete_one({"accession_number": str(accession_number)})
                    if result.deleted_count > 0:
                        # Set flash message for success
                        flash(f"Book with accession number {accession_number} deleted successfully!", "success")
                        return jsonify({"success": True, "message": "Book deleted successfully."})
            
            return jsonify({"success": False, "message": "Book not found."})
        
        # If we found the book with the original accession number, delete it
        result = books_collection.delete_one({"accession_number": accession_number})
        
        if result.deleted_count > 0:
            # Set flash message for success
            flash(f"Book with accession number {accession_number} deleted successfully!", "success")
            return jsonify({"success": True, "message": "Book deleted successfully."})
        else:
            return jsonify({"success": False, "message": "Book not found."})

    except Exception as e:
        print("Error deleting book:", str(e))
        return jsonify({"success": False, "message": str(e)})

@app.route("/delete_book_success/<accession_number>")
def delete_book_success(accession_number):
    flash("Book deleted successfully.", "success")
    return redirect(url_for("all_books"))

@app.route('/single-date-report', methods=['GET', 'POST'])
def single_date_report():
    report_data = None
    selected_date = None
    if request.method == 'POST':
        selected_date = request.form.get('report_date')
        if selected_date:
            # Find all issued_books with issued_at matching selected_date
            report_data = list(issued_books.find({
                'issued_at': {'$regex': f'^{selected_date}'}
            }))
    return render_template('single_date_report.html', report_data=report_data, selected_date=selected_date)

# Download Excel for Single Date Report
@app.route('/download-single-date-report', methods=['POST'])
def download_single_date_report():
    selected_date = request.form.get('report_date')
    if not selected_date:
        flash('No date selected', 'error')
        return redirect(url_for('single_date_report'))
    report_data = list(issued_books.find({
        'issued_at': {'$regex': f'^{selected_date}'}
    }))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Single Date Report'
    headers = ['Student Name', 'Roll No', 'Class', 'Book Title', 'Author', 'Department', 'Barcode', 'Issued Date', 'Returned Date', 'Status']
    ws.append(headers)
    for book in report_data:
        ws.append([
            book['student'].get('studentName', ''),
            book['student'].get('rollno', ''),
            book['student'].get('section', ''),
            book['book'].get('title', ''),
            book['book'].get('author', ''),
            book['book'].get('department', ''),
            book['book'].get('barcode', ''),
            book.get('issued_at', ''),
            book.get('returned_at', ''),
            book.get('status', '')
        ])
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename=Single_Date_Report_{selected_date}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

# Range Date Report Page
@app.route('/range-date-report', methods=['GET', 'POST'])
def range_date_report():
    report_data = None
    start_date = None
    end_date = None
    if request.method == 'POST':
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        if start_date and end_date:
            # Find all issued_books with issued_at between start_date and end_date (inclusive)
            report_data = list(issued_books.find({
                'issued_at': {
                    '$gte': start_date,
                    '$lte': end_date + ' 23:59:59'
                }
            }))
    return render_template('range_date_report.html', report_data=report_data, start_date=start_date, end_date=end_date)

# Download Excel for Range Date Report
@app.route('/download-range-date-report', methods=['POST'])
def download_range_date_report():
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    if not start_date or not end_date:
        flash('Start and end date required', 'error')
        return redirect(url_for('range_date_report'))
    report_data = list(issued_books.find({
        'issued_at': {
            '$gte': start_date,
            '$lte': end_date + ' 23:59:59'
        }
    }))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Date Range Report'
    headers = ['Student Name', 'Roll No', 'Class', 'Book Title', 'Author', 'Department', 'Barcode', 'Issued Date', 'Status']
    ws.append(headers)
    for book in report_data:
        ws.append([
            book['student'].get('studentName', ''),
            book['student'].get('rollno', ''),
            book['student'].get('section', ''),
            book['book'].get('title', ''),
            book['book'].get('author', ''),
            book['book'].get('department', ''),
            book['book'].get('barcode', ''),
            book.get('issued_at', ''),
            book.get('status', '')
        ])
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename=Date_Range_Report_{start_date}_to_{end_date}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


if __name__ == '__main__':
    app.debug = True
    app.run(host='0.0.0.0', port=5000)

